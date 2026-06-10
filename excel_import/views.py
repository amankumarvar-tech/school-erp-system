from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
import io, datetime

def parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.datetime.strptime(str(val).strip(), fmt).date()
        except:
            pass
    return None

def str_val(v):
    if v is None:
        return ''
    return str(v).strip()

@login_required
def import_home(request):
    return render(request, 'excel_import/home.html')

@login_required
def import_students(request):
    if request.method != 'POST':
        return render(request, 'excel_import/import_students.html')

    f = request.FILES.get('excel_file')
    if not f:
        messages.error(request, 'Please select an Excel file.')
        return render(request, 'excel_import/import_students.html')

    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        # Support both "Students" sheet and first sheet
        ws = wb['Students'] if 'Students' in wb.sheetnames else wb.active
    except Exception as e:
        messages.error(request, f'Could not read Excel file: {e}')
        return render(request, 'excel_import/import_students.html')

    from student_mgmt.models import Student, Class
    from core.models import AcademicYear

    # Find header row (row with "Admission No" or "admission")
    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and 'admission' in str(cell.value).lower():
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        messages.error(request, 'Header row not found. Make sure your Excel has correct column headers.')
        return render(request, 'excel_import/import_students.html')

    # Map headers to column indices
    col_map = {}
    for cell in ws[header_row]:
        if cell.value:
            key = str(cell.value).lower().replace('\n', ' ').strip()
            col_map[key] = cell.column - 1  # 0-indexed

    def get_col(row_vals, *keys):
        for k in keys:
            for map_key, idx in col_map.items():
                if k in map_key:
                    if idx < len(row_vals):
                        return str_val(row_vals[idx])
        return ''

    academic_year = AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.first()

    success, skipped, errors = 0, 0, []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row):
            continue
        row = list(row)

        adm_no = get_col(row, 'admission no', 'admission number')
        first_name = get_col(row, 'first name')
        last_name = get_col(row, 'last name')
        dob_raw = get_col(row, 'date of birth', 'dob')
        gender = get_col(row, 'gender')
        class_name = get_col(row, 'class name', 'class')
        section = get_col(row, 'section')
        father_name = get_col(row, 'father name')
        father_phone = get_col(row, 'father phone')
        mother_name = get_col(row, 'mother name')
        mother_phone = get_col(row, 'mother phone')
        address = get_col(row, 'address')
        email = get_col(row, 'email')
        phone = get_col(row, 'phone')
        blood_group = get_col(row, 'blood group')
        category = get_col(row, 'category')
        adm_date_raw = get_col(row, 'admission date')
        roll_number = get_col(row, 'roll number')

        if not adm_no or not first_name:
            skipped += 1
            continue

        # Skip example rows
        if adm_no.upper().startswith('STU00') and first_name.lower() in ['rahul', 'priya', 'akash']:
            # Allow if they want to keep example data — skip only truly empty
            pass

        dob = parse_date(dob_raw)
        adm_date = parse_date(adm_date_raw) or datetime.date.today()

        if not dob:
            errors.append(f"Row {adm_no}: Invalid date of birth '{dob_raw}'")
            skipped += 1
            continue

        # Normalize gender
        g = gender.upper()[0] if gender else 'M'
        if g not in ['M', 'F', 'O']:
            g = 'M'

        # Normalize category
        cat = category.upper() if category else 'GEN'
        if cat not in ['GEN', 'OBC', 'SC', 'ST', 'EWS']:
            cat = 'GEN'

        # Normalize blood group
        bg = blood_group.upper().replace(' ', '') if blood_group else ''
        valid_bg = ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-']
        if bg not in valid_bg:
            bg = ''

        # Get or create class
        student_class = None
        if class_name and section and academic_year:
            student_class, _ = Class.objects.get_or_create(
                name=class_name, section=section.upper(), academic_year=academic_year,
                defaults={'capacity': 40}
            )

        try:
            obj, created = Student.objects.update_or_create(
                admission_number=adm_no,
                defaults={
                    'first_name': first_name,
                    'last_name': last_name,
                    'date_of_birth': dob,
                    'gender': g,
                    'father_name': father_name or 'N/A',
                    'father_phone': father_phone or '0000000000',
                    'mother_name': mother_name or '',
                    'mother_phone': mother_phone or '',
                    'address': address or 'N/A',
                    'email': email or '',
                    'phone': phone or '',
                    'blood_group': bg,
                    'category': cat,
                    'admission_date': adm_date,
                    'roll_number': roll_number or '',
                    'current_class': student_class,
                    'is_active': True,
                }
            )
            success += 1
        except Exception as e:
            errors.append(f"Row {adm_no}: {str(e)[:80]}")
            skipped += 1

    msg = f"✅ {success} students imported/updated."
    if skipped:
        msg += f" ⚠️ {skipped} rows skipped."
    messages.success(request, msg)
    for err in errors[:5]:
        messages.warning(request, err)

    return render(request, 'excel_import/import_students.html', {
        'success': success, 'skipped': skipped, 'errors': errors
    })


@login_required
def import_teachers(request):
    if request.method != 'POST':
        return render(request, 'excel_import/import_teachers.html')

    f = request.FILES.get('excel_file')
    if not f:
        messages.error(request, 'Please select an Excel file.')
        return render(request, 'excel_import/import_teachers.html')

    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb['Teachers'] if 'Teachers' in wb.sheetnames else wb.active
    except Exception as e:
        messages.error(request, f'Could not read file: {e}')
        return render(request, 'excel_import/import_teachers.html')

    from teacher_mgmt.models import Teacher

    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and 'employee' in str(cell.value).lower():
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        messages.error(request, 'Header row not found.')
        return render(request, 'excel_import/import_teachers.html')

    col_map = {}
    for cell in ws[header_row]:
        if cell.value:
            col_map[str(cell.value).lower().replace('\n',' ').strip()] = cell.column - 1

    def get_col(row_vals, *keys):
        for k in keys:
            for map_key, idx in col_map.items():
                if k in map_key:
                    if idx < len(row_vals):
                        return str_val(row_vals[idx])
        return ''

    success, skipped, errors = 0, 0, []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row):
            continue
        row = list(row)

        emp_id = get_col(row, 'employee id')
        first_name = get_col(row, 'first name')
        last_name = get_col(row, 'last name')
        dob_raw = get_col(row, 'date of birth', 'dob')
        gender = get_col(row, 'gender')
        email = get_col(row, 'email')
        phone = get_col(row, 'phone')
        address = get_col(row, 'address')
        designation = get_col(row, 'designation')
        qualification = get_col(row, 'qualification')
        exp = get_col(row, 'experience')
        joining_raw = get_col(row, 'joining date')
        emp_type = get_col(row, 'employment type')
        salary = get_col(row, 'basic salary', 'salary')

        if not emp_id or not first_name:
            skipped += 1
            continue

        dob = parse_date(dob_raw)
        joining = parse_date(joining_raw) or datetime.date.today()

        if not dob:
            errors.append(f"Row {emp_id}: Invalid DOB '{dob_raw}'")
            skipped += 1
            continue

        g = (gender.upper()[0] if gender else 'M')
        if g not in ['M', 'F', 'O']:
            g = 'M'

        emp_type_val = emp_type.lower() if emp_type else 'permanent'
        if emp_type_val not in ['permanent', 'contract', 'part_time', 'guest']:
            emp_type_val = 'permanent'

        try:
            sal = float(salary) if salary else 0
        except:
            sal = 0

        try:
            exp_int = int(float(exp)) if exp else 0
        except:
            exp_int = 0

        try:
            Teacher.objects.update_or_create(
                employee_id=emp_id,
                defaults={
                    'first_name': first_name,
                    'last_name': last_name,
                    'date_of_birth': dob,
                    'gender': g,
                    'email': email or f'{emp_id.lower()}@school.edu',
                    'phone': phone or '0000000000',
                    'address': address or 'N/A',
                    'designation': designation or 'Teacher',
                    'qualification': qualification or 'B.Ed',
                    'experience_years': exp_int,
                    'joining_date': joining,
                    'employment_type': emp_type_val,
                    'basic_salary': sal,
                    'is_active': True,
                }
            )
            success += 1
        except Exception as e:
            errors.append(f"Row {emp_id}: {str(e)[:80]}")
            skipped += 1

    msg = f"✅ {success} teachers imported/updated."
    if skipped:
        msg += f" ⚠️ {skipped} rows skipped."
    messages.success(request, msg)
    for err in errors[:5]:
        messages.warning(request, err)

    return render(request, 'excel_import/import_teachers.html', {
        'success': success, 'skipped': skipped, 'errors': errors
    })


@login_required
def import_fees(request):
    if request.method != 'POST':
        return render(request, 'excel_import/import_fees.html')

    f = request.FILES.get('excel_file')
    if not f:
        messages.error(request, 'Please select an Excel file.')
        return render(request, 'excel_import/import_fees.html')

    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb['Fees'] if 'Fees' in wb.sheetnames else wb.active
    except Exception as e:
        messages.error(request, f'Could not read file: {e}')
        return render(request, 'excel_import/import_fees.html')

    from fees_mgmt.models import FeePayment, FeeStructure
    from student_mgmt.models import Student

    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and 'admission' in str(cell.value).lower():
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        messages.error(request, 'Header row not found.')
        return render(request, 'excel_import/import_fees.html')

    col_map = {}
    for cell in ws[header_row]:
        if cell.value:
            col_map[str(cell.value).lower().replace('\n',' ').strip()] = cell.column - 1

    def get_col(row_vals, *keys):
        for k in keys:
            for map_key, idx in col_map.items():
                if k in map_key:
                    if idx < len(row_vals):
                        return str_val(row_vals[idx])
        return ''

    success, skipped, errors = 0, 0, []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row):
            continue
        row = list(row)

        adm_no = get_col(row, 'admission no')
        fee_type = get_col(row, 'fee type')
        total_amount = get_col(row, 'total amount')
        due_date_raw = get_col(row, 'due date')
        amount_paid = get_col(row, 'amount paid')
        payment_date_raw = get_col(row, 'payment date')
        pay_method = get_col(row, 'payment method')
        remarks = get_col(row, 'remarks')

        if not adm_no or not fee_type or not total_amount:
            skipped += 1
            continue

        try:
            student = Student.objects.get(admission_number=adm_no)
        except Student.DoesNotExist:
            errors.append(f"Student '{adm_no}' not found — import students first.")
            skipped += 1
            continue

        due_date = parse_date(due_date_raw) or datetime.date.today()
        payment_date = parse_date(payment_date_raw)

        try:
            total = float(total_amount)
            paid = float(amount_paid) if amount_paid else 0
        except:
            skipped += 1
            continue

        fee_type_val = fee_type.lower().strip()
        valid_fee_types = ['tuition', 'exam', 'transport', 'hostel', 'library', 'other', 'admission', 'sports', 'computer']
        if fee_type_val not in valid_fee_types:
            fee_type_val = 'other'

        pay_method_val = pay_method.lower().strip() if pay_method else 'cash'
        valid_pay = ['cash', 'bank', 'cheque', 'online', 'upi', 'dd']
        if pay_method_val not in valid_pay:
            pay_method_val = 'cash'

        if paid >= total:
            status = 'paid'
        elif paid > 0:
            status = 'partial'
        else:
            status = 'pending'
            if due_date < datetime.date.today():
                status = 'overdue'

        # Get or create fee structure
        fee_struct, _ = FeeStructure.objects.get_or_create(
            student_class=student.current_class,
            fee_type=fee_type_val,
            defaults={'amount': total, 'academic_year': student.current_class.academic_year if student.current_class else None}
        ) if student.current_class else (None, False)

        import uuid
        receipt_no = f"REC-{datetime.date.today().year}-{str(uuid.uuid4())[:6].upper()}"

        try:
            FeePayment.objects.create(
                student=student,
                fee_structure=fee_struct,
                total_amount=total,
                amount_paid=paid,
                due_date=due_date,
                payment_date=payment_date,
                payment_method=pay_method_val,
                status=status,
                receipt_number=receipt_no,
                remarks=remarks or '',
            )
            success += 1
        except Exception as e:
            errors.append(f"Row {adm_no}: {str(e)[:80]}")
            skipped += 1

    msg = f"✅ {success} fee records imported."
    if skipped:
        msg += f" ⚠️ {skipped} rows skipped."
    messages.success(request, msg)
    for err in errors[:5]:
        messages.warning(request, err)

    return render(request, 'excel_import/import_fees.html', {
        'success': success, 'skipped': skipped, 'errors': errors
    })


@login_required
def download_template(request):
    """Serve the template Excel file"""
    import os
    from django.conf import settings

    # Generate template on the fly
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        def sh(cell, bg="1a1f36", fg="FFFFFF"):
            cell.font = Font(bold=True, color=fg, size=10, name="Arial")
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def se(cell):
            cell.font = Font(color="666666", size=9, name="Arial", italic=True)
            cell.fill = PatternFill("solid", start_color="F7F8FC")

        def tb():
            s = Side(style='thin', color='CCCCCC')
            return Border(left=s, right=s, top=s, bottom=s)

        wb = openpyxl.Workbook()

        # Students sheet
        ws = wb.active
        ws.title = "Students"
        ws.merge_cells("A1:R1")
        ws["A1"] = "📚 STUDENT IMPORT TEMPLATE — School ERP"
        ws["A1"].font = Font(bold=True, size=13, color="1a1f36", name="Arial")
        ws["A1"].fill = PatternFill("solid", start_color="E8ECFF")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.merge_cells("A2:R2")
        ws["A2"] = "⚠️ Instructions: Do NOT change headers. Date format: YYYY-MM-DD. Required fields marked with *. Rows 4-6 are EXAMPLES — replace with real data."
        ws["A2"].font = Font(size=9, color="991B1B", name="Arial")
        ws["A2"].fill = PatternFill("solid", start_color="FEE2E2")
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 18

        h = [("Admission No *",18),("First Name *",15),("Last Name *",15),("Date of Birth *\n(YYYY-MM-DD)",18),("Gender *\n(M/F/O)",12),("Class Name *",12),("Section *",10),("Father Name *",18),("Father Phone *",14),("Mother Name",18),("Mother Phone",14),("Address *",22),("Email",20),("Phone",13),("Blood Group",14),("Category\n(GEN/OBC/SC/ST/EWS)",16),("Admission Date *\n(YYYY-MM-DD)",18),("Roll Number",12)]
        for i,(lbl,w) in enumerate(h,1):
            c = ws.cell(row=3,column=i,value=lbl)
            sh(c); c.border = tb()
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[3].height = 36

        ex = [
            ["STU001","Rahul","Sharma","2010-06-15","M","10","A","Rajesh Sharma","9876543210","Priya Sharma","9876543211","123 Main St, Delhi","","","B+","GEN","2024-04-01","1"],
            ["STU002","Priya","Verma","2011-03-22","F","9","B","Amit Verma","9812345678","Sunita Verma","9812345679","456 Park Ave, Mumbai","","","O+","OBC","2024-04-01","2"],
            ["STU003","Akash","Singh","2009-11-10","M","11","A","Vikram Singh","9765432109","","","789 Rose Garden, Jaipur","","","A+","SC","2023-04-01","3"],
        ]
        for ri, rd in enumerate(ex,4):
            for ci, v in enumerate(rd,1):
                c = ws.cell(row=ri,column=ci,value=v); se(c); c.border = tb()
        for r in range(7,57):
            for c in range(1,19):
                cell = ws.cell(row=r,column=c,value=""); cell.border = tb()
        ws.freeze_panes = "A4"

        # Teachers sheet
        ws2 = wb.create_sheet("Teachers")
        ws2.merge_cells("A1:N1")
        ws2["A1"] = "👨‍🏫 TEACHER IMPORT TEMPLATE — School ERP"
        ws2["A1"].font = Font(bold=True, size=13, color="1a1f36", name="Arial")
        ws2["A1"].fill = PatternFill("solid", start_color="E8ECFF")
        ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 28
        ws2.merge_cells("A2:N2")
        ws2["A2"] = "⚠️ Instructions: Do NOT change headers. Date format: YYYY-MM-DD. Rows 4-6 are EXAMPLES — replace with real data."
        ws2["A2"].font = Font(size=9, color="991B1B", name="Arial")
        ws2["A2"].fill = PatternFill("solid", start_color="FEE2E2")
        ws2["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[2].height = 18

        th = [("Employee ID *",15),("First Name *",15),("Last Name *",15),("Date of Birth *\n(YYYY-MM-DD)",18),("Gender *\n(M/F/O)",12),("Email *",22),("Phone *",14),("Address *",22),("Designation *",18),("Qualification *",18),("Experience Yrs",14),("Joining Date *\n(YYYY-MM-DD)",18),("Employment Type\n(permanent/contract/part_time)",22),("Basic Salary",14)]
        for i,(lbl,w) in enumerate(th,1):
            c = ws2.cell(row=3,column=i,value=lbl); sh(c,bg="0D3349"); c.border = tb()
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.row_dimensions[3].height = 36

        tex = [
            ["EMP001","Sunita","Mishra","1985-04-10","F","sunita@school.edu","9876543000","45 Teacher Colony Delhi","Principal","M.Ed, B.Ed",15,"2010-06-01","permanent",55000],
            ["EMP002","Ramesh","Kumar","1990-08-22","M","ramesh@school.edu","9812340000","12 Staff Quarters Delhi","Math Teacher","M.Sc",8,"2016-07-01","permanent",38000],
            ["EMP003","Anita","Joshi","1992-12-05","F","anita@school.edu","9765430000","78 Sector 5 Noida","Science Teacher","M.Sc, B.Ed",5,"2019-04-01","contract",32000],
        ]
        for ri,rd in enumerate(tex,4):
            for ci,v in enumerate(rd,1):
                c = ws2.cell(row=ri,column=ci,value=v); se(c); c.border = tb()
        for r in range(7,57):
            for c in range(1,15):
                cell = ws2.cell(row=r,column=c,value=""); cell.border = tb()
        ws2.freeze_panes = "A4"

        # Fees sheet
        ws3 = wb.create_sheet("Fees")
        ws3.merge_cells("A1:H1")
        ws3["A1"] = "💰 FEE IMPORT TEMPLATE — School ERP"
        ws3["A1"].font = Font(bold=True, size=13, color="1a1f36", name="Arial")
        ws3["A1"].fill = PatternFill("solid", start_color="E8ECFF")
        ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[1].height = 28
        ws3.merge_cells("A2:H2")
        ws3["A2"] = "⚠️ Import students FIRST before importing fees. Admission No must match."
        ws3["A2"].font = Font(size=9, color="991B1B", name="Arial")
        ws3["A2"].fill = PatternFill("solid", start_color="FEE2E2")
        ws3["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws3.row_dimensions[2].height = 18

        fh = [("Admission No *",16),("Fee Type *\n(tuition/exam/transport etc)",22),("Total Amount *",15),("Due Date *\n(YYYY-MM-DD)",18),("Amount Paid",14),("Payment Date\n(YYYY-MM-DD)",18),("Payment Method\n(cash/bank/upi/cheque)",20),("Remarks",20)]
        for i,(lbl,w) in enumerate(fh,1):
            c = ws3.cell(row=3,column=i,value=lbl); sh(c,bg="065F46"); c.border = tb()
            ws3.column_dimensions[get_column_letter(i)].width = w
        ws3.row_dimensions[3].height = 36

        fex = [
            ["STU001","tuition",15000,"2024-04-30",15000,"2024-04-15","bank","April fee"],
            ["STU001","exam",2000,"2024-05-10",0,"","",""],
            ["STU002","tuition",15000,"2024-04-30",8000,"2024-04-20","cash","Partial"],
        ]
        for ri,rd in enumerate(fex,4):
            for ci,v in enumerate(rd,1):
                c = ws3.cell(row=ri,column=ci,value=v); se(c); c.border = tb()
        for r in range(7,57):
            for c in range(1,9):
                cell = ws3.cell(row=r,column=c,value=""); cell.border = tb()
        ws3.freeze_panes = "A4"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="SchoolERP_Import_Template.xlsx"'
        return response

    except Exception as e:
        return HttpResponse(f"Error generating template: {e}", status=500)
